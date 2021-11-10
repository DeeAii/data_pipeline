import openpyxl
import re
import textblob
import autocorrect
import string
import pandas as pd
from openpyxl import Workbook


inputPath = 'D:\\SERS\\5103 Indegenious\\'
inputFile = 'Submit+Performance+Commitments+#+1+by+11_59+p.m.xlsx'

# wb_obj = openpyxl.load_workbook("D:\\SERS\\Topic modelling\\Qualtrix\\Takeaways_August 30, 2021_15.33.xlsx")

wb_obj = openpyxl.load_workbook(inputPath+inputFile)
sheet_obj = wb_obj.active

# cell_obj = sheet_obj.cell(row=1,column=1)

# sheet = wb_obj['Sheet0']
max_row = sheet_obj.max_row
max_col = sheet_obj.max_column

if sheet_obj.cell(row=1,column=1).value.upper() == "RECORDEDDATE":
    myList = []
    for i in range(3, max_row + 1):
        for j in range(2, max_col+1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            if cell_obj.value is not None:
                myList.append(cell_obj.value.capitalize())
else:
    myList = []
    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=8)
        if cell_obj.value is not None and cell_obj.value != '':
            myList.append(cell_obj.value.capitalize())

print(myList)
print(len(myList))
myList.sort()
print(myList)

def remove_punc(string):
    punc = '''!()[]{};:'"\,<>./?@#$%^&*_~'''
    for ele in string:
        if ele in punc:
            string = string.replace(ele,"")
    return string

myList = [remove_punc(i) for i in myList]
print(myList)

def remove_duplicate(x):
    return list(dict.fromkeys(x))

myList = remove_duplicate(myList)

print(myList)
print(len(myList))

myWorkbook = openpyxl.Workbook()
mySheet = myWorkbook.active

mySheet["A1"] = "name"
mySheet["B1"] = "content"

outputPath = 'D:\\SERS\\5103 Indegenious\\'
outCSVPath = 'D:/SERS/5103 Indegenious/'
outputExcel = 'Processed takeaways.xlsx'
outputCSV = 'Processed takeaways.csv'

myWorkbook.save(outputPath+outputExcel)

# spell check with autocorrect
from autocorrect import Speller
check = Speller(lang='en')

cList = []

for i in range(len(myList)):
    txt = myList[i]
    corrected = check(txt)
    cList.append(corrected)

print(cList)

j = 0
for i in range(2,len(myList)+2):
    cellref = mySheet.cell(row=i,column=2)
    cellref.value = cList[j]
    j = j+1


k = 000
# print(k)
for i in range(2, len(myList)+2):
    cellref = mySheet.cell(row=i, column=1)
    k = k+1
    cellref.value = "A"+str(k)

myWorkbook.save(outputPath+outputExcel)

read_file = pd.read_excel(outCSVPath+outputExcel,sheet_name='Sheet')
read_file.to_csv (outCSVPath+outputCSV, encoding='utf-8', index=None, header=True)

